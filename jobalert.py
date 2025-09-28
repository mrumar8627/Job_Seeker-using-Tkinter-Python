import requests
from bs4 import BeautifulSoup
import time
import datetime
import os
from openpyxl import Workbook, load_workbook
import tkinter as tk
import webbrowser
import threading
import sys
import pystray
from PIL import Image

# ---------------- Resource Path Helper ----------------
def resource_path(relative_path):
    """Get absolute path to resource, works for dev and PyInstaller EXE"""
    try:
        base_path = sys._MEIPASS  # PyInstaller temp folder
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# ---------------- Excel Setup ----------------
FILE_NAME = resource_path("govt_jobs.xlsx")

def save_job_to_excel(title, link, source):
    """Save job details to Excel file"""
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date Found", "Job Title", "Link", "Source", "Status"])
        wb.save(FILE_NAME)

    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
               title, link, source, "Not Applied"])
    wb.save(FILE_NAME)

# ---------------- Popup Setup ----------------
snooze_time = 3600  # 1 hour

def show_popup(title, link):
    """Fullscreen popup with Apply and Snooze buttons"""
    root = tk.Tk()
    root.title("🚨 Government Job Alert 🚨")
    root.attributes("-fullscreen", True)

    tk.Label(root, text=f"New Job: {title}", font=("Arial", 30, "bold"), fg="red").pack(pady=40)
    tk.Label(root, text=link, font=("Arial", 18), fg="blue").pack(pady=20)

    def open_link():
        webbrowser.open(link)
        root.destroy()

    def snooze():
        root.destroy()
        threading.Timer(snooze_time, lambda: show_popup(title, link)).start()

    tk.Button(root, text="✅ Apply Now", font=("Arial", 20), bg="green", command=open_link).pack(pady=20)
    tk.Button(root, text="⏰ Snooze 1 Hour", font=("Arial", 20), bg="orange", command=snooze).pack(pady=20)

    root.mainloop()

# ---------------- Job Sources ----------------
JOB_SOURCES = {
    "FPSC": "https://fpsc.gov.pk/jobs/gr/currentjobs",
    "PPSC": "https://www.ppsc.gop.pk/Jobs.aspx",
    "NTS": "https://www.nts.org.pk/new/Allresults.php"
}

# ---------------- Memory ----------------
seen_jobs = set()

# ---------------- Scraper ----------------
def check_new_jobs():
    """Check all job sources for new jobs"""
    for source, url in JOB_SOURCES.items():
        try:
            response = requests.get(url, timeout=10)
            soup = BeautifulSoup(response.text, "html.parser")

            if source == "PPSC":
                for row in soup.find_all("tr"):
                    cols = row.find_all("td")
                    if len(cols) >= 2:
                        title = cols[1].get_text(strip=True)
                        link_tag = cols[1].find("a", href=True)
                        if title and link_tag:
                            href = link_tag["href"]
                            job_link = f"https://www.ppsc.gop.pk/{href}"

                            if title not in seen_jobs:
                                seen_jobs.add(title)
                                save_job_to_excel(title, job_link, source)
                                show_popup(title, job_link)

            else:
                for link in soup.find_all("a", href=True):
                    title = link.text.strip()
                    href = link["href"]

                    if "job" in href.lower() and title != "":
                        from urllib.parse import urljoin
                        job_link = urljoin(url, href)

                        if title not in seen_jobs:
                            seen_jobs.add(title)
                            save_job_to_excel(title, job_link, source)
                            show_popup(title, job_link)

        except Exception as e:
            print(f"❌ Error checking {source}: {e}")

# ---------------- Background Job Loop ----------------
def job_loop():
    while True:
        check_new_jobs()
        time.sleep(300)  # check every 5 min

# ---------------- System Tray ----------------
ICON_FILE = resource_path("icon.png")

def on_quit(icon, item):
    icon.stop()
    os._exit(0)  # hard exit so threads die

def run_tray():
    # Try to load icon, else fallback red square
    try:
        if os.path.exists(ICON_FILE):
            image = Image.open(ICON_FILE)
        else:
            raise FileNotFoundError
    except Exception:
        image = Image.new('RGB', (64, 64), color=(255, 0, 0))  # fallback red square

    menu = pystray.Menu(
        pystray.MenuItem("Quit", on_quit)
    )

    icon = pystray.Icon("Govt Job Alert", image, "Govt Job Alert", menu)

    # run job loop in background
    threading.Thread(target=job_loop, daemon=True).start()

    # start system tray
    icon.run()

# ---------------- Main ----------------
if __name__ == "__main__":
    run_tray()
